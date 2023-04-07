from openpyxl import load_workbook 

ws= None
wb= None

def LoadWorkbook(name):
    global ws
    global wb
    ws= load_workbook(name)
    wb= ws.active()

    return
``
def insertRow(rowData):
    ws.append(rowData)

def insertCol(rowno,colno,props):
    for val in props:
        ws.cell(rowno,colno).value= val
        colno =colno+1

    return




def ReadSheet(){
    for row in ws.iter_rows(min_row=2):
    col1_val = row[0].value
    name = col1_val.split("_")
    namecode =
    props = getAdmetProps(col1_val)
    new_col = row[2].column + 1  
    insertCols(row[0].row,new_col,props)
}




LoadWorkbook("./ADMET_Properties2.xlsx")
wb.save("./ADMET_Properties2.xlsx")
